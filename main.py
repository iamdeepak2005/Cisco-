import re
import ipaddress
import networkx as nx
import matplotlib.pyplot as plt
from pathlib import Path
from collections import Counter
import itertools
from faults import simulate_link_failure
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
def parse_config(file_path):
    """Extract hostname, interfaces, and protocols from a router config file."""
    data = {"hostname": None, "interfaces": [], "protocols": []}

    with open(file_path, "r") as f:
        content = f.read()

    # hostname
    hostname_match = re.search(r"hostname\s+(\S+)", content)
    if hostname_match:
        data["hostname"] = hostname_match.group(1)

    # interfaces with IP and bandwidth
    interfaces = re.findall(
        r"interface\s+(\S+)\s+ip address\s+(\S+)\s+(\S+)(?:.*?bandwidth\s+(\d+))?(?:.*?mtu\s+(\d+))?",
        content, re.S)
    for intf, ip, mask, bw, mtu in interfaces:
        data["interfaces"].append({
            "name": intf,
            "ip": ip,
            "mask": mask,
            "bandwidth": int(bw) if bw else 100000,  # default if missing
            "mtu": int(mtu) if mtu else 1500        # default if missing
        })

    # routing protocols
    if "router ospf" in content:
        data["protocols"].append("OSPF")
    if "router bgp" in content:
        data["protocols"].append("BGP")

    return data


def same_subnet(ip1, mask1, ip2, mask2):
    """Check if two interfaces are in the same subnet."""
    net1 = ipaddress.ip_network(f"{ip1}/{mask1}", strict=False)
    net2 = ipaddress.ip_network(f"{ip2}/{mask2}", strict=False)
    return net1 == net2

def check_mtu_mismatches(G, router_data):
    issues = []
    hostname_map = {r["hostname"]: r for r in router_data}

    for u, v in G.edges():
        r1 = hostname_map[u]
        r2 = hostname_map[v]

        # Compare MTU on interfaces in same subnet
        for intf1 in r1["interfaces"]:
            for intf2 in r2["interfaces"]:
                net1 = f"{intf1['ip']}/{intf1['mask']}"
                net2 = f"{intf2['ip']}/{intf2['mask']}"
                try:
                    import ipaddress
                    if ipaddress.ip_network(net1, strict=False) == ipaddress.ip_network(net2, strict=False):
                        if intf1["mtu"] != intf2["mtu"]:
                            issues.append(f"MTU mismatch on link {u}–{v}: {intf1['mtu']} vs {intf2['mtu']}")
                except:
                    pass
    return issues

def build_topology(router_data):
    """Construct a graph of routers using networkx."""
    G = nx.Graph()

    # Add routers as nodes
    for r in router_data:
        G.add_node(r["hostname"], protocols=r["protocols"])

    # Add edges if interfaces share subnet
    for i, r1 in enumerate(router_data):
        for j, r2 in enumerate(router_data):
            if i >= j:  # avoid duplicates
                continue
            for intf1 in r1["interfaces"]:
                for intf2 in r2["interfaces"]:
                    if same_subnet(intf1["ip"], intf1["mask"], intf2["ip"], intf2["mask"]):
                        bw = min(intf1["bandwidth"], intf2["bandwidth"])
                        G.add_edge(r1["hostname"], r2["hostname"], bandwidth=bw)

    return G



# Load an icon image
def get_icon(path, zoom=0.1):
    img = plt.imread(path)
    return OffsetImage(img, zoom=zoom)

def visualize_topology_with_icons(G, utilization=None):
    # Fixed positions: layered network
    pos = {
        "PC2": (-2, 0),
        "FW1": (-1, 0),
        "R1": (0, 0),
        "R2": (1, 0),
        "R3": (2, 0),
        "SW1": (2, -1),
        "PC1": (3, 0)
    }

    fig, ax = plt.subplots(figsize=(10,6))

    # Edge colors by utilization
    edge_colors = []
    edge_labels = {}
    for u, v, d in G.edges(data=True):
        color = "black"
        cap = d.get("capacity_kbps", 0)
        load = utilization.get(frozenset((u, v)), 0) if utilization else 0
        util = (load / cap) if cap else 0

        if utilization:
            if util > 1.0:
                color = "red"
            elif util > 0.7:
                color = "orange"
            else:
                color = "green"

            # Label includes both capacity and % utilization
            edge_labels[(u, v)] = f"{cap//1000} Mbps / {util:.0%}"
        else:
            edge_labels[(u, v)] = f"{cap//1000} Mbps"

        edge_colors.append(color)

    # Draw edges
    nx.draw_networkx_edges(G, pos, ax=ax, edge_color=edge_colors, width=2)

    # Draw labels with capacity + utilization
    nx.draw_networkx_edge_labels(G, pos, edge_labels=edge_labels, font_size=8, rotate=False)

    # Place icons (same as before)
    for n in G.nodes():
        if n.startswith("R"):
            icon = get_icon("icons/router.png", zoom=0.1)
        elif n.startswith("S"):
            icon = get_icon("icons/switch.png", zoom=0.1)
        elif n.startswith("PC"):
            icon = get_icon("icons/pc.png", zoom=0.07)
        elif n.startswith("FW"):
            icon = get_icon("icons/firewall.png", zoom=0.1)
        else:
            icon = get_icon("icons/device.png", zoom=0.1)

        ab = AnnotationBbox(icon, pos[n], frameon=False)
        ax.add_artist(ab)

    # Node labels under icons
    for n, (x, y) in pos.items():
        ax.text(x, y-0.15, n, ha="center", va="top", fontsize=9, fontweight="bold")

    plt.axis("off")
    plt.tight_layout()
    plt.show()

def validate_network(router_data, G):
    issues = []
    # duplicate IPs
    all_ips = [intf["ip"] for r in router_data for intf in r["interfaces"]]
    for ip, count in Counter(all_ips).items():
        if count > 1:
            issues.append(f"Duplicate IP detected: {ip}")

    # loops
    cycles = list(nx.cycle_basis(G))
    if cycles:
        issues.append(f"Network loops found: {cycles}")

    # protocol recommendation
    ospf_count = sum(1 for r in router_data if "OSPF" in r["protocols"])
    if ospf_count == len(router_data) and len(G.nodes) > 5:
        issues.append("Recommendation: Use BGP instead of OSPF for scalability.")

    # MTU mismatches
    issues.extend(check_mtu_mismatches(G, router_data))

    return issues


def rename_capacity(G):
    # Normalize edge attr name
    for u, v, d in G.edges(data=True):
        cap = d.get("bandwidth") or d.get("capacity_kbps")
        d["capacity_kbps"] = int(cap)

def path_edges(path):
    return list(zip(path[:-1], path[1:]))

def simulate_traffic(G, demands_mbps):
    """
    demands_mbps: list of (src_router, dst_router, mbps)
    Returns: edge_loads_kbps dict keyed by frozenset({u,v})
    """
    rename_capacity(G)
    edge_loads = {frozenset((u, v)): 0 for u, v in G.edges()}
    for s, t, mbps in demands_mbps:
        if s == t:
            continue
        if not nx.has_path(G, s, t):
            print(f"[WARN] No path between {s} and {t}. Demand {mbps} Mbps dropped.")
            continue
        # Shortest path by hop count; you can change to weighted later
        path = nx.shortest_path(G, s, t)
        kbps = int(mbps * 1000)
        for u, v in path_edges(path):
            edge_loads[frozenset((u, v))] += kbps
    return edge_loads

def utilization_report(G, edge_loads):
    rows = []
    for u, v, d in G.edges(data=True):
        cap = d["capacity_kbps"]
        load = edge_loads[frozenset((u, v))]
        util = load / cap if cap else float("inf")
        rows.append((u, v, cap, load, util))
    # Sort by highest utilization
    rows.sort(key=lambda x: x[4], reverse=True)
    print("\n=== Link Utilization ===")
    for u, v, cap, load, util in rows:
        status = "OK"
        if util > 1.0:
            status = "OVERLOADED"
        elif util > 0.7:
            status = "HIGH"
        print(f"{u:<6}—{v:<6}  cap={cap:>8} kbps  load={load:>8} kbps  util={util:6.2%}  {status}")
    return rows

def recommend_load_balancing(G, demands_mbps, edge_loads, max_k=3):
    """
    For overloaded edges, suggest secondary paths for the specific flows causing congestion.
    """
    print("\n=== Load Balancing Recommendations ===")
    overloaded_edges = {
        frozenset((u, v))
        for u, v, d in G.edges(data=True)
        if edge_loads[frozenset((u, v))] > d["capacity_kbps"]
    }
    if not overloaded_edges:
        print("No overloaded links. No action needed.")
        return

    # Build quick index of path per demand to see which flows touch overloaded edges
    for s, t, mbps in demands_mbps:
        if not nx.has_path(G, s, t):
            continue
        primary = nx.shortest_path(G, s, t)
        primary_edges = {frozenset(e) for e in path_edges(primary)}
        if primary_edges & overloaded_edges:
            # Try alternate simple paths
            tried = 0
            suggested = None
            for path in nx.shortest_simple_paths(G, s, t):
                if path == primary:
                    continue
                tried += 1
                alt_edges = [frozenset(e) for e in path_edges(path)]
                # Check if alternate reduces pressure (prefer edge-disjoint with any overloaded edge)
                if not (set(alt_edges) & overloaded_edges):
                    suggested = path
                    break
                if tried >= max_k:
                    break
            if suggested:
                print(f"{s}->{t} ({mbps} Mbps): activate secondary path {suggested} for lower-priority traffic.")
            else:
                print(f"{s}->{t} ({mbps} Mbps): no clean alternate found. Consider capacity upgrade or policy-based split.")

def run_capacity_check(G, demands_mbps):
    edge_loads = simulate_traffic(G, demands_mbps)
    utilization_report(G, edge_loads)
    recommend_load_balancing(G, demands_mbps, edge_loads)
    return edge_loads

def utilization_dataframe(G, edge_loads):
    rows = []
    for u, v, d in G.edges(data=True):
        cap = int(d.get("capacity_kbps") or d.get("bandwidth") or 0)
        load = int(edge_loads.get(frozenset((u, v)), 0))
        util = load / cap if cap else 0
        rows.append({"u": u, "v": v, "capacity_kbps": cap, "load_kbps": load, "util": util})
    df = pd.DataFrame(rows).sort_values("util", ascending=False)
    return df

def write_excel_report(df: pd.DataFrame, path="reports/utilization.xlsx"):
    Path("reports").mkdir(exist_ok=True, parents=True)
    df2 = df.copy()
    df2["util_pct"] = (df2["util"] * 100).round(1)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df2.to_excel(writer, index=False, sheet_name="links")
    wb = load_workbook(path)
    ws = wb["links"]
    # Apply 3-color scale to util_pct column (E is util, F is util_pct after export; adjust if columns shift)
    # Find the header row to locate the column:
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    util_col = headers.get("util_pct", "F")
    last_row = ws.max_row
    rng = f"{util_col}2:{util_col}{last_row}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(start_type="num", start_value=0, start_color="63BE7B",  # green
                       mid_type="num", mid_value=70, mid_color="FFC000",      # orange
                       end_type="num", end_value=100, end_color="FF0000")     # red
    )
    wb.save(path)
    print(f"[INFO] Excel written: {path}")
if __name__ == "__main__":
    conf_dir = Path("Conf")
    router_data = []
    for file in conf_dir.glob("*/config.dump"):
        parsed = parse_config(file)
        router_data.append(parsed)

    G = build_topology(router_data)
    rename_capacity(G)   # normalize attributes

    # Validation
    problems = validate_network(router_data, G)
    print("\nValidation Report:")
    if problems:
        for p in problems:
            print(" -", p)
    else:
        print("No major issues found.")

    # Define traffic demands
    demands = [
        ("PC1", "PC2", 20),   # user-to-user traffic
        ("R1", "R3", 50),     # inter-router
        ("PC1", "R1", 10),    # host to gateway
        ("R2", "FW1", 15),    # router to firewall
        ("R3", "PC2", 25),    # router to host
    ]

    # Run baseline utilization check
    edge_loads = run_capacity_check(G, demands)
    df = utilization_dataframe(G, edge_loads)
    write_excel_report(df, path="reports/utilization.xlsx")

    # 🔥 Draw topology with utilization coloring
    visualize_topology_with_icons(G, utilization=edge_loads)

    # Simulate a link failure (e.g., R2–R3 goes down)
    simulate_link_failure(G.copy(), ("R2", "R3"), demands)

    # Extra validation
    problems = validate_network(router_data, G)
    mtu_issues = check_mtu_mismatches(G, router_data)
    problems.extend(mtu_issues)

    print("\nValidation Report:")
    if problems:
        for p in problems:
            print(" -", p)
    else:
        print("No major issues found.")
